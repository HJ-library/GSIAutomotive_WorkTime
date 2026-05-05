import re

with open('logic.py', 'r', encoding='utf-8') as f:
    content = f.read()

export_old = '''def export_monthly_excel(user_id, month_str, filepath):
    if not user_id:
        return False
        
    try:
        conn = get_connection()
        df = pd.read_sql_query('SELECT * FROM work_logs WHERE user_id = ? AND date LIKE ? ORDER BY date', conn, params=(user_id, month_str + '%',))
        conn.close()
        
        if df.empty:
            print(f"Export failed: No data for {user_id} in {month_str}")
            return False
            
        # Ensure numeric columns are actually numeric and handle NaNs
        numeric_cols = ['non_work_time', 'break_time', 'work_time', 'overtime_earned', 'overtime_used']
        for col in numeric_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
            
        def mins_to_hm(mins):
            if pd.isna(mins) or mins == 0:
                return "00:00"
            mins = int(mins)
            h = mins // 60
            m = mins % 60
            return f"{h:02d}:{m:02d}"

        def mins_to_hours(mins):
            if pd.isna(mins) or mins == 0:
                return None
            return round(int(mins) / 60, 2)

        df['비근무시간'] = df['non_work_time'].apply(mins_to_hm)
        df['휴게시간'] = df['break_time'].apply(mins_to_hm)
        df['실근무시간'] = df['work_time'].apply(mins_to_hm)
        df['발생연장근로'] = df['overtime_earned'].apply(mins_to_hours)
        df['사용연장근로'] = df['overtime_used'].apply(mins_to_hours)
        
        # Formatting work type for excel
        def format_type(row):
            t_names = {
                "normal": "정상근무", 
                "morning_half": "오전 반차", 
                "afternoon_half": "오후 반차", 
                "holiday": "공휴일",
                "annual_leave": "연차",
                "public_leave": "공가",
                "sick_leave": "병가"
            }
            base_name = t_names.get(row['type'], row['type'])
            if row['type'] == 'holiday' and row['description']:
                return f"{base_name} - {row['description']}"
            if row['description'] and row['type'] != 'holiday':
                return f"{base_name} ({row['description']})"
            return base_name

        df['근무형태'] = df.apply(format_type, axis=1)

        # Exact columns from reference file
        cols = ['date', '근무형태', 'clock_in', 'clock_out', '비근무시간', '휴게시간', '실근무시간', '발생연장근로', '사용연장근로']
        
        out_df = df[cols].copy()
        out_df.columns = ['일자', '근무형태', '출근', '퇴근', '비근무시간', '휴게시간', '실근무시간', '발생연장근로', '사용연장근로']
        
        # Save basic dataframe to excel first
        out_df.to_excel(filepath, index=False)
        
        # Post-process with openpyxl to add styles and summary block
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            
            wb = load_workbook(filepath)
            ws = wb.active
            
            yellow_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
            bold_font = Font(bold=True)
            center_align = Alignment(horizontal="center", vertical="center")
            
            # Format header
            for col_idx in range(1, len(out_df.columns) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = bold_font
                cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                cell.alignment = center_align
                
            # openpyxl is 1-indexed. Row 1 is header. Data starts at row 2.
            for r_idx, row in enumerate(df.itertuples(), start=2):
                if getattr(row, 'overtime_earned', 0) > 0:
                    for c_idx in range(1, len(out_df.columns) + 1):
                        ws.cell(row=r_idx, column=c_idx).fill = yellow_fill
                        
            # Add summary block on the right (like the reference file)
            stats = get_monthly_summary(user_id, month_str)
            
            ws.cell(row=2, column=11, value="총 발생 연장근로").font = bold_font
            ws.cell(row=2, column=12, value=round(stats['curr_earned'] / 60, 2))
            
            ws.cell(row=3, column=11, value="총 사용 연장근로").font = bold_font
            ws.cell(row=3, column=12, value=round(stats['curr_used'] / 60, 2))
            
            ws.cell(row=4, column=11, value="이월된 연장근로").font = bold_font
            ws.cell(row=4, column=12, value=round(stats['carry_over'] / 60, 2))
            
            ws.cell(row=5, column=11, value="잔여 연장근로 시간").font = bold_font
            ws.cell(row=5, column=12, value=round(stats['total_remaining'] / 60, 2)).font = bold_font
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
                
            wb.save(filepath)
        except Exception as style_e:
            print(f"Excel styling failed (continuing): {style_e}")
            
        return True
    except Exception as e:
        print(f"Logic export error: {e}")
        raise e'''

export_new = '''def export_monthly_excel(user_id, month_str, filepath):
    if not user_id:
        return False
        
    try:
        conn = get_connection()
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute('SELECT * FROM work_logs WHERE user_id = ? AND date LIKE ? ORDER BY date', (user_id, month_str + '%',))
        rows = c.fetchall()
        conn.close()
        
        if not rows:
            print(f"Export failed: No data for {user_id} in {month_str}")
            return False
            
        def mins_to_hm(mins):
            if not mins or mins == 0: return "00:00"
            mins = int(mins)
            return f"{mins//60:02d}:{mins%60:02d}"

        def mins_to_hours(mins):
            if not mins or mins == 0: return None
            return round(int(mins) / 60, 2)

        def format_type(row):
            t_names = {
                "normal": "정상근무", "morning_half": "오전 반차", "afternoon_half": "오후 반차", 
                "holiday": "공휴일", "annual_leave": "연차", "public_leave": "공가", "sick_leave": "병가"
            }
            base_name = t_names.get(row['type'], row['type'])
            desc = row['description']
            if row['type'] == 'holiday' and desc:
                return f"{base_name} - {desc}"
            if desc and row['type'] != 'holiday':
                return f"{base_name} ({desc})"
            return base_name

        wb = Workbook()
        ws = wb.active
        
        headers = ['일자', '근무형태', '출근', '퇴근', '비근무시간', '휴게시간', '실근무시간', '발생연장근로', '사용연장근로']
        ws.append(headers)
        
        for r in rows:
            ws.append([
                r['date'],
                format_type(r),
                r['clock_in'],
                r['clock_out'],
                mins_to_hm(r['non_work_time']),
                mins_to_hm(r['break_time']),
                mins_to_hm(r['work_time']),
                mins_to_hours(r['overtime_earned']),
                mins_to_hours(r['overtime_used'])
            ])
            
        # Post-process with openpyxl to add styles and summary block
        try:
            yellow_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
            bold_font = Font(bold=True)
            center_align = Alignment(horizontal="center", vertical="center")
            
            # Format header
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = bold_font
                cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                cell.alignment = center_align
                
            for r_idx, r in enumerate(rows, start=2):
                if (r['overtime_earned'] or 0) > 0:
                    for c_idx in range(1, len(headers) + 1):
                        ws.cell(row=r_idx, column=c_idx).fill = yellow_fill
                        
            stats = get_monthly_summary(user_id, month_str)
            
            ws.cell(row=2, column=11, value="총 발생 연장근로").font = bold_font
            ws.cell(row=2, column=12, value=round(stats['curr_earned'] / 60, 2))
            
            ws.cell(row=3, column=11, value="총 사용 연장근로").font = bold_font
            ws.cell(row=3, column=12, value=round(stats['curr_used'] / 60, 2))
            
            ws.cell(row=4, column=11, value="이월된 연장근로").font = bold_font
            ws.cell(row=4, column=12, value=round(stats['carry_over'] / 60, 2))
            
            ws.cell(row=5, column=11, value="잔여 연장근로 시간").font = bold_font
            ws.cell(row=5, column=12, value=round(stats['total_remaining'] / 60, 2)).font = bold_font
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column].width = max_length + 2
                
            wb.save(filepath)
        except Exception as style_e:
            print(f"Excel styling failed (continuing): {style_e}")
            wb.save(filepath)
            
        return True
    except Exception as e:
        print(f"Logic export error: {e}")
        raise e'''

content = content.replace(export_old, export_new)
with open('logic.py', 'w', encoding='utf-8') as f:
    f.write(content)
print('Phase 2 done')
