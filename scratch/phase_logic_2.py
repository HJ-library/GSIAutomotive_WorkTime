import sys

with open('logic.py', 'r', encoding='utf-8') as f:
    content = f.read()

new_logic = '''
def auto_fill_missing_days(user_id, month_str):
    if not user_id: return 0
    from datetime import datetime, date, timedelta
    import calendar
    import sqlite3
    
    year, month = map(int, month_str.split('-'))
    num_days = calendar.monthrange(year, month)[1]
    
    start_date = date(year, month, 1)
    end_date = min(date(year, month, num_days), date.today() - timedelta(days=1))
    
    conn = get_connection()
    c = conn.cursor()
    c.execute('SELECT date FROM work_logs WHERE user_id = ? AND date LIKE ?', (user_id, month_str + '%'))
    existing_dates = set([row[0] for row in c.fetchall()])
    conn.close()
    
    curr = start_date
    count = 0
    while curr <= end_date:
        d_str = curr.strftime("%Y-%m-%d")
        if curr.weekday() < 5 and d_str not in existing_dates: # 평일이고 기록이 없으면
            save_work_log(user_id, d_str, 'normal', '08:30', '17:30', 0, 0, '')
            count += 1
        curr += timedelta(days=1)
    return count

def export_yearly_excel(year_str, target_user_id, filepath):
    import sqlite3
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    
    conn = get_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    if target_user_id == 'all':
        c.execute('SELECT id, name FROM users ORDER BY name')
        users = c.fetchall()
    else:
        c.execute('SELECT id, name FROM users WHERE id = ?', (target_user_id,))
        users = c.fetchall()
        
    wb = Workbook()
    if not users:
        wb.save(filepath)
        conn.close()
        return True
        
    first_sheet = True
    
    for u in users:
        u_id = u['id']
        u_name = u['name']
        
        c.execute('SELECT * FROM work_logs WHERE user_id = ? AND date LIKE ? ORDER BY date', (u_id, year_str + '%'))
        rows = c.fetchall()
        
        if first_sheet:
            ws = wb.active
            ws.title = u_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=u_name)
            
        def mins_to_hm(mins):
            if not mins or mins == 0: return "00:00"
            mins = int(mins)
            return f"{mins//60:02d}:{mins%60:02d}"

        def mins_to_hours(mins):
            if not mins or mins == 0: return None
            return round(int(mins) / 60, 2)

        def format_type(row):
            t_names = {
                "normal": "정상근무", "morning_half": "오전 반차", "afternoon_half": "오후 반차", 
                "holiday": "공휴일 / 특별휴가", "annual_leave": "연차", "public_leave": "공가", "sick_leave": "병가",
                "replacement_leave": "휴무", "public_leave_half": "반공가 + 반차"
            }
            base_name = t_names.get(row['type'], row['type'])
            desc = row['description']
            if row['type'] == 'holiday' and desc: return f"{base_name} - {desc}"
            if desc and row['type'] != 'holiday': return f"{base_name} ({desc})"
            return base_name

        headers = ['일자', '근무형태', '출근', '퇴근', '비근무시간', '휴게시간', '실근무시간', '발생연장근로', '사용연장근로']
        ws.append(headers)
        
        for r in rows:
            ws.append([
                r['date'], format_type(r), r['clock_in'], r['clock_out'],
                mins_to_hm(r['non_work_time']), mins_to_hm(r['break_time']), mins_to_hm(r['work_time']),
                mins_to_hours(r['overtime_earned']), mins_to_hours(r['overtime_used'])
            ])
            
        try:
            yellow_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
            bold_font = Font(bold=True)
            center_align = Alignment(horizontal="center", vertical="center")
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = bold_font
                cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                cell.alignment = center_align
                
            for r_idx, r in enumerate(rows, start=2):
                if (r['overtime_earned'] or 0) > 0:
                    for c_idx in range(1, len(headers) + 1):
                        ws.cell(row=r_idx, column=c_idx).fill = yellow_fill
                        
            # Yearly summary
            c.execute("SELECT SUM(overtime_earned) as e, SUM(overtime_used) as u FROM work_logs WHERE user_id = ? AND date LIKE ?", (u_id, year_str + '%'))
            yearly_stats = c.fetchone()
            y_earned = yearly_stats['e'] if yearly_stats['e'] else 0
            y_used = yearly_stats['u'] if yearly_stats['u'] else 0
            
            ws.cell(row=2, column=11, value="년도 총 발생 연장근로").font = bold_font
            ws.cell(row=2, column=12, value=round(y_earned / 60, 2))
            
            ws.cell(row=3, column=11, value="년도 총 사용 연장근로").font = bold_font
            ws.cell(row=3, column=12, value=round(y_used / 60, 2))
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column].width = max_length + 2
        except Exception as style_e:
            pass

    wb.save(filepath)
    conn.close()
    return True
'''

content += new_logic

with open('logic.py', 'w', encoding='utf-8') as f:
    f.write(content)
