import math
import calendar
import sqlite3
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from database import get_connection

def get_users():
    conn = get_connection()
    c = conn.cursor()
    c.execute('SELECT id, name FROM users ORDER BY name ASC')
    users = c.fetchall()
    conn.close()
    return [{"id": u[0], "name": u[1]} for u in users]

def add_user(name):
    conn = get_connection()
    c = conn.cursor()
    try:
        c.execute('INSERT INTO users (name) VALUES (?)', (name,))
        conn.commit()
        user_id = c.lastrowid
        status = "success"
        msg = "User added"
    except sqlite3.IntegrityError:
        user_id = None
        status = "error"
        msg = "User already exists"
    conn.close()
    return {"status": status, "user_id": user_id, "message": msg}

def delete_user(user_id):
    conn = get_connection()
    c = conn.cursor()
    c.execute('DELETE FROM work_logs WHERE user_id = ?', (user_id,))
    c.execute('DELETE FROM users WHERE id = ?', (user_id,))
    conn.commit()
    conn.close()
    return {"status": "success", "message": "User deleted"}

def change_admin_password(old_pw, new_pw):
    conn = get_connection()
    c = conn.cursor()
    c.execute('SELECT password FROM admin_settings WHERE id = 1')
    row = c.fetchone()
    if not row or row[0] != old_pw:
        conn.close()
        return False, "기존 비밀번호가 일치하지 않습니다."
        
    c.execute('UPDATE admin_settings SET password = ? WHERE id = 1', (new_pw,))
    conn.commit()
    conn.close()
    return True, "비밀번호가 성공적으로 변경되었습니다."

def get_overtime_usage_history(user_id, year_str):
    if not user_id:
        return []
    conn = get_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute('''
        SELECT date, type, overtime_used, description 
        FROM work_logs 
        WHERE user_id = ? AND date LIKE ? AND overtime_used > 0
        ORDER BY date DESC
    ''', (user_id, year_str + '%'))
    rows = c.fetchall()
    conn.close()
    return [dict(row) for row in rows]

def calculate_work_time(clock_in_str, clock_out_str, non_work_minutes):
    if not clock_in_str or not clock_out_str:
        return 0, 0
        
    fmt = "%H:%M"
    try:
        t_in = datetime.strptime(clock_in_str, fmt)
        t_out = datetime.strptime(clock_out_str, fmt)
    except ValueError:
        return 0, 0
        
    if t_out < t_in:
        t_out += timedelta(days=1)
        
    total_duration_minutes = int((t_out - t_in).total_seconds() / 60)
    target_duration = total_duration_minutes - non_work_minutes
    
    if target_duration > 540:
        base_work = 480
        base_break = 60
        extra_presence = target_duration - 540
        extra_break = 0
        extra_work = 0
        
        if extra_presence > 0:
            b = min(30, extra_presence)
            extra_break += b
            extra_presence -= b
            
        if extra_presence > 0:
            w = min(480, extra_presence)
            extra_work += w
            extra_presence -= w
            
        while extra_presence > 0:
            b = min(30, extra_presence)
            extra_break += b
            extra_presence -= b
            if extra_presence > 0:
                w = min(240, extra_presence)
                extra_work += w
                extra_presence -= w
                
        return base_work + extra_work, base_break + extra_break
    else:
        W = target_duration
        breaks = 0
        while W >= 240 * (breaks + 1):
            W -= 30
            breaks += 1
        return W, breaks * 30

def save_work_log(user_id, date_str, log_type, clock_in, clock_out, non_work_time, overtime_used=0, description=""):
    if not user_id:
        return
        
    conn = get_connection()
    c = conn.cursor()
    
    if log_type in ['holiday', 'annual_leave', 'public_leave', 'sick_leave', 'replacement_leave', 'public_leave_half']:
        work_minutes = 8 * 60
        break_minutes = 0
        clock_in = '08:30'
        clock_out = '17:30'
        non_work_time = 0
        if log_type == 'replacement_leave' and overtime_used == 0:
            overtime_used = 480
    elif log_type in ['morning_half', 'afternoon_half']:
        work_minutes = 4 * 60
        break_minutes = 0
        non_work_time = 0
        if log_type == 'morning_half':
            clock_in = '08:30'
            clock_out = '12:30'
        else:
            clock_in = '13:30'
            clock_out = '17:30'
    else:
        work_minutes, break_minutes = calculate_work_time(clock_in, clock_out, non_work_time)
        
    try:
        c.execute('''
            INSERT OR REPLACE INTO work_logs 
            (user_id, date, type, clock_in, clock_out, non_work_time, break_time, work_time, overtime_earned, overtime_used, description)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
        ''', (user_id, date_str, log_type, clock_in, clock_out, non_work_time, break_minutes, work_minutes, overtime_used, description))
        conn.commit()
    except Exception as e:
        print(f"Error saving log: {e}")
        raise e
    finally:
        conn.close()
    
    recalculate_monthly_overtime(user_id, date_str[:7])

def recalculate_monthly_overtime(user_id, month_str):
    try:
        conn = get_connection()
        c = conn.cursor()
        
        c.execute('SELECT date, work_time, type FROM work_logs WHERE user_id = ? AND date LIKE ? ORDER BY date', (user_id, month_str + '%'))
        logs = c.fetchall()
        
        accumulated_raw_overtime = 0
        weekly_overtime = {}
        
        for row in logs:
            date_str = row[0]
            work_time = row[1]
            log_type = row[2]
            
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            is_weekend = dt.weekday() >= 5
            year, week_num, _ = dt.isocalendar()
            week_key = (year, week_num)
            
            if week_key not in weekly_overtime:
                weekly_overtime[week_key] = 0
            
            raw_earned = 0
            if is_weekend:
                raw_earned = work_time
            else:
                if work_time > 480:
                    raw_earned = work_time - 480
                    
            # 30분 단위로 내림(버림) 처리
            raw_earned = (raw_earned // 30) * 30
            
            weekly_overtime[week_key] += raw_earned
                    
            earned = 0
            if raw_earned > 0:
                available_1x = max(0, 1200 - accumulated_raw_overtime)
                
                if raw_earned <= available_1x:
                    earned = raw_earned
                else:
                    portion_1x = available_1x
                    portion_15x = raw_earned - available_1x
                    earned = portion_1x + int(portion_15x * 1.5)
                    
                accumulated_raw_overtime += raw_earned
                
            c.execute('UPDATE work_logs SET overtime_earned = ? WHERE user_id = ? AND date = ?', (earned, user_id, date_str))
            
        conn.commit()
    except Exception as e:
        print(f"Error recalculating overtime: {e}")
    finally:
        conn.close()

def get_work_logs(user_id, start_date_str, end_date_str):
    if not user_id:
        return []
    conn = get_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute('SELECT * FROM work_logs WHERE user_id = ? AND date >= ? AND date <= ? ORDER BY date', (user_id, start_date_str, end_date_str))
    rows = c.fetchall()
    conn.close()
    return [dict(row) for row in rows]

def add_months(sourcedate, months):
    month = sourcedate.month - 1 + months
    year = sourcedate.year + month // 12
    month = month % 12 + 1
    day = min(sourcedate.day, calendar.monthrange(year, month)[1])
    return datetime(year, month, day).date()

def get_monthly_summary(user_id, month_str):
    if not user_id:
        return {"carry_over": 0, "curr_earned": 0, "curr_used": 0, "expired_overtime": 0, "expired_by_date": {}, "expiring_soon": 0, "total_remaining": 0}
        
    import calendar
    from datetime import datetime
    year, month = map(int, month_str.split('-'))
    last_day = calendar.monthrange(year, month)[1]
    
    start_date = f"{month_str}-01"
    end_date = f"{month_str}-{last_day:02d}"
    
    conn = get_connection()
    c = conn.cursor()
    
    c.execute('SELECT date, overtime_earned, overtime_used FROM work_logs WHERE user_id = ? AND date <= ? ORDER BY date', (user_id, end_date))
    logs = c.fetchall()
    
    c.execute('SELECT date, is_extended, expires_at FROM overtime_vault WHERE user_id = ?', (user_id,))
    existing_vault = {row[0]: (row[1], row[2]) for row in c.fetchall()}
    
    vault = []
    carry_over = -1
    
    for row in logs:
        dt_str = row[0]
        earned = row[1]
        used = row[2]
        
        if dt_str >= start_date and carry_over == -1:
            carry_over = sum(v['earned'] - v['used'] for v in vault if v['expires_at'] >= start_date)
            
        if earned > 0:
            if dt_str in existing_vault:
                is_extended, expires_at = existing_vault[dt_str]
            else:
                dt = datetime.strptime(dt_str, "%Y-%m-%d").date()
                expires_at = add_months(dt, 1).strftime("%Y-%m-%d")
            vault.append({'date': dt_str, 'earned': earned, 'used': 0, 'expires_at': expires_at})
            
        if used > 0:
            rem = used
            valid_vaults = [v for v in vault if v['earned'] > v['used'] and v['expires_at'] >= dt_str]
            for v in valid_vaults:
                if rem <= 0: break
                avail = v['earned'] - v['used']
                ded = min(avail, rem)
                v['used'] += ded
                rem -= ded

    if carry_over == -1:
        carry_over = sum(v['earned'] - v['used'] for v in vault if v['expires_at'] >= start_date)
        
    total_remaining = sum(v['earned'] - v['used'] for v in vault if v['expires_at'] > end_date)
    
    c.execute('SELECT SUM(overtime_earned), SUM(overtime_used) FROM work_logs WHERE user_id = ? AND date LIKE ?', (user_id, month_str + '%'))
    r2 = c.fetchone()
    curr_earned = r2[0] or 0
    curr_used = r2[1] or 0
    
    expired_by_date = {}
    for v in vault:
        rem = v['earned'] - v['used']
        if rem > 0 and start_date <= v['expires_at'] <= end_date:
            expired_by_date[v['expires_at']] = expired_by_date.get(v['expires_at'], 0) + rem

    expired_overtime = carry_over + curr_earned - curr_used - total_remaining
    if expired_overtime < 0:
        expired_overtime = 0

    conn.close()
    
    return {
        "carry_over": carry_over,
        "curr_earned": curr_earned,
        "curr_used": curr_used,
        "expired_overtime": expired_overtime,
        "expired_by_date": expired_by_date,
        "expiring_soon": 0,
        "total_remaining": total_remaining
    }

def get_weekly_work_time(user_id, start_date_str, end_date_str):
    if not user_id:
        return 0
    conn = get_connection()
    c = conn.cursor()
    c.execute('SELECT SUM(work_time) FROM work_logs WHERE user_id = ? AND date >= ? AND date <= ?', (user_id, start_date_str, end_date_str))
    row = c.fetchone()
    conn.close()
    return row[0] or 0

def export_monthly_excel(user_id, month_str, filepath):
    if not user_id:
        return False
        
    try:
        conn = get_connection()
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute('SELECT * FROM work_logs WHERE user_id = ? AND date LIKE ? ORDER BY date', (user_id, month_str + '%',))
        rows = c.fetchall()
        conn.close()
        
        if not rows:
            print(f"Export failed: No data for {user_id} in {month_str}")
            return False
            
        def mins_to_hm(mins):
            if not mins or mins == 0: return "00:00"
            mins = int(mins)
            return f"{mins//60:02d}:{mins%60:02d}"

        def mins_to_hours(mins):
            if not mins or mins == 0: return None
            return round(int(mins) / 60, 2)

        def format_type(row):
            t_names = {
                "normal": "정상근무", "morning_half": "오전 반차", "afternoon_half": "오후 반차", 
                "holiday": "공휴일 / 특별휴가", "annual_leave": "연차", "public_leave": "공가", "sick_leave": "병가",
                "replacement_leave": "휴무", "public_leave_half": "반공가 + 반차"
            }
            base_name = t_names.get(row['type'], row['type'])
            desc = row['description']
            if row['type'] == 'holiday' and desc:
                return f"{base_name} - {desc}"
            if desc and row['type'] != 'holiday':
                return f"{base_name} ({desc})"
            return base_name

        wb = Workbook()
        ws = wb.active
        
        stats = get_monthly_summary(user_id, month_str)
        current_carry_over = stats['carry_over']
        expired_by_date = stats.get('expired_by_date', {})
        
        headers = ['일자', '근무형태', '출근', '퇴근', '비근무시간', '휴게시간', '실근무시간', '발생연장근로', '사용연장근로', '소멸연장근로', '잔여연장근로']
        ws.append(headers)
        
        last_dt = f"{month_str}-00"
        actual_expired_total = 0
        
        for r in rows:
            earned = r['overtime_earned'] or 0
            used = r['overtime_used'] or 0
            dt_str = r['date']
            
            expired = sum(v for k, v in expired_by_date.items() if last_dt < k <= dt_str)
            actual_expired_total += expired
            last_dt = dt_str
            
            current_carry_over = current_carry_over + earned - used - expired
            
            ws.append([
                dt_str,
                format_type(r),
                r['clock_in'],
                r['clock_out'],
                mins_to_hm(r['non_work_time']),
                mins_to_hm(r['break_time']),
                mins_to_hm(r['work_time']),
                mins_to_hours(earned),
                mins_to_hours(used),
                mins_to_hours(expired),
                mins_to_hours(current_carry_over)
            ])
            
        try:
            yellow_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
            bold_font = Font(bold=True)
            red_font = Font(color="EF4444", bold=True)
            center_align = Alignment(horizontal="center", vertical="center")
            
            text_red = Font(color="FF0000")
            text_blue = Font(color="0000FF")
            text_green = Font(color="008000")
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = bold_font
                cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                cell.alignment = center_align
                
            for r_idx, r in enumerate(rows, start=2):
                if (r['overtime_earned'] or 0) > 0:
                    for c_idx in range(1, len(headers) + 1):
                        ws.cell(row=r_idx, column=c_idx).fill = yellow_fill
                        
                t = r['type']
                if t in ['holiday', 'public_leave', 'annual_leave', 'sick_leave']:
                    ws.cell(row=r_idx, column=2).font = text_red
                elif t in ['replacement_leave']:
                    ws.cell(row=r_idx, column=2).font = text_blue
                elif t in ['morning_half', 'afternoon_half', 'public_leave_half']:
                    ws.cell(row=r_idx, column=2).font = text_green
                        
            ws.cell(row=2, column=13, value="이월된 연장근로").font = bold_font
            ws.cell(row=2, column=14, value=round(stats['carry_over'] / 60, 2))
            
            ws.cell(row=3, column=13, value="총 발생 연장근로").font = bold_font
            ws.cell(row=3, column=14, value=round(stats['curr_earned'] / 60, 2))
            
            ws.cell(row=4, column=13, value="총 사용 연장근로").font = bold_font
            ws.cell(row=4, column=14, value=round(stats['curr_used'] / 60, 2))
            
            ws.cell(row=5, column=13, value="소멸된 연장근로").font = bold_font
            ws.cell(row=5, column=14, value=round(actual_expired_total / 60, 2)).font = red_font
            
            ws.cell(row=6, column=13, value="잔여 연장근로 시간").font = bold_font
            ws.cell(row=6, column=14, value=round(current_carry_over / 60, 2)).font = bold_font
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column].width = max_length + 2
                
            wb.save(filepath)
        except Exception as style_e:
            print(f"Excel styling failed (continuing): {style_e}")
            wb.save(filepath)
            
        return True
    except Exception as e:
        print(f"Logic export error: {e}")
        raise e

def get_yearly_overtime_summary(user_id):
    if not user_id:
        return []
        
    conn = get_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("""
        SELECT strftime('%Y-%m', date) as month, SUM(overtime_earned) as total
        FROM work_logs
        WHERE user_id = ? AND date >= date('now', '-1 year')
        GROUP BY month
        ORDER BY month DESC
    """, (user_id,))
    rows = c.fetchall()
    conn.close()
    return [dict(row) for row in rows]

def generate_import_template(filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "근무기록_양식"
    
    headers = ['일자', '근무형태', '출근', '퇴근', '비근무시간', '휴게시간', '실근무시간', '발생연장근로', '사용연장근로']
    ws.append(headers)
    
    ws.append(['2026-04-01', '정상 근무', '08:30', '17:30', '00:00', '00:00', '08:00', None, None])
    ws.append(['2026-04-02', '공휴일 / 특별휴가 (현충일)', '08:30', '17:30', '00:00', '00:00', '08:00', None, None])
    
    wb.save(filepath)
    return True

def import_excel(user_id, filepath):
    if not user_id:
        return False, "사용자를 선택해주세요."
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows: return False, "데이터가 없습니다."
        
        headers = rows[0]
        try:
            date_idx = headers.index('일자')
            type_idx = headers.index('근무형태')
            in_idx = headers.index('출근')
            out_idx = headers.index('퇴근')
            nw_idx = headers.index('비근무시간')
            use_idx = headers.index('사용연장근로')
        except ValueError:
            return False, "양식이 올바르지 않습니다."
            
        def hm_to_mins(hm_str):
            if not hm_str: return 0
            try:
                parts = str(hm_str).strip().split(':')
                return int(parts[0])*60 + int(parts[1])
            except:
                return 0
                
        for row in rows[1:]:
            date_val = str(row[date_idx])[:10]
            if date_val == 'None' or not date_val.strip(): continue
            
            type_str = str(row[type_idx] or '').strip()
            desc = ""
            base_type = "normal"
            
            if "반공가" in type_str: base_type = "public_leave_half"
            elif "반차" in type_str:
                base_type = "morning_half" if "오전" in type_str else "afternoon_half"
            elif "연차" in type_str: base_type = "annual_leave"
            elif "공가" in type_str: base_type = "public_leave"
            elif "병가" in type_str: base_type = "sick_leave"
            elif "휴무" in type_str or "연장" in type_str or "replacement" in type_str: base_type = "replacement_leave"
            elif "공휴일" in type_str or "특별휴가" in type_str or "기념일" in type_str:
                base_type = "holiday"
            
            if "(" in type_str and ")" in type_str:
                desc = type_str.split("(")[1].split(")")[0]
            elif " - " in type_str:
                desc = type_str.split(" - ")[1]
            elif base_type == "holiday" and type_str != "공휴일":
                desc = type_str
                
            non_work = hm_to_mins(row[nw_idx])
            
            used_hours = row[use_idx] if row[use_idx] is not None else 0
            try:
                used_mins = int(float(used_hours) * 60)
                used_mins = (used_mins // 30) * 30
            except:
                used_mins = 0
            
            cin = str(row[in_idx] or '').strip()
            cout = str(row[out_idx] or '').strip()
            if cin == 'None': cin = ''
            if cout == 'None': cout = ''
            
            save_work_log(
                user_id,
                date_val,
                base_type,
                cin,
                cout,
                non_work,
                used_mins,
                desc
            )
        return True, "성공적으로 불러왔습니다."
    except Exception as e:
        return False, str(e)

def sync_vault(user_id):
    if not user_id: return
    conn = get_connection()
    c = conn.cursor()
    
    # Reset used_minutes
    c.execute('UPDATE overtime_vault SET used_minutes = 0 WHERE user_id = ?', (user_id,))
    
    c.execute('SELECT date, overtime_earned, overtime_used FROM work_logs WHERE user_id = ? ORDER BY date', (user_id,))
    logs = c.fetchall()
    
    for row in logs:
        dt_str = row[0]
        earned = row[1]
        used = row[2]
        
        if earned > 0:
            c.execute('SELECT id FROM overtime_vault WHERE user_id = ? AND date = ?', (user_id, dt_str))
            if not c.fetchone():
                dt = datetime.strptime(dt_str, "%Y-%m-%d").date()
                exp_at = add_months(dt, 1).strftime("%Y-%m-%d")
                c.execute('INSERT INTO overtime_vault (user_id, date, earned_minutes, used_minutes, expires_at, is_extended) VALUES (?, ?, ?, 0, ?, 0)', (user_id, dt_str, earned, exp_at))
            else:
                c.execute('UPDATE overtime_vault SET earned_minutes = ? WHERE user_id = ? AND date = ?', (earned, user_id, dt_str))
                
        if used > 0:
            rem = used
            c.execute('SELECT id, earned_minutes, used_minutes FROM overtime_vault WHERE user_id = ? AND earned_minutes > used_minutes AND expires_at >= ? ORDER BY date', (user_id, dt_str))
            valid_vaults = c.fetchall()
            for v in valid_vaults:
                if rem <= 0: break
                avail = v[1] - v[2]
                ded = min(avail, rem)
                c.execute('UPDATE overtime_vault SET used_minutes = used_minutes + ? WHERE id = ?', (ded, v[0]))
                rem -= ded

    conn.commit()
    conn.close()

def get_yearly_ledger(user_id, year_str):
    if not user_id:
        return []
    sync_vault(user_id)
        
    conn = get_connection()
    c = conn.cursor()
    
    # Get all logs for user up to end of year
    c.execute('SELECT date, overtime_earned, overtime_used FROM work_logs WHERE user_id = ? ORDER BY date', (user_id,))
    logs = c.fetchall()
    
    # Get all vaults for user
    c.execute('SELECT id, date, earned_minutes, used_minutes, expires_at, is_extended FROM overtime_vault WHERE user_id = ?', (user_id,))
    vaults = c.fetchall()
    
    conn.close()
    
    expirations_by_date = {}
    expiring_vault_by_date = {}
    vault_by_date = {}
    for v in vaults:
        v_id, v_date, earned, used, expires_at, is_ext = v
        rem = earned - used
        if rem > 0:
            expirations_by_date[expires_at] = expirations_by_date.get(expires_at, 0) + rem
            if expires_at not in expiring_vault_by_date:
                expiring_vault_by_date[expires_at] = v
        vault_by_date[v_date] = v
        
    dates_set = set([row[0] for row in logs])
    dates_set.update(expirations_by_date.keys())
    
    all_dates = sorted(list(dates_set))
    
    ledger = []
    running_balance = 0
    
    for dt_str in all_dates:
        log = next((r for r in logs if r[0] == dt_str), None)
        earned = log[1] if log else 0
        used = log[2] if log else 0
        expired = expirations_by_date.get(dt_str, 0)
        
        running_balance = running_balance + earned - used - expired
        
        if dt_str.startswith(year_str) and (earned > 0 or used > 0 or expired > 0):
            v = vault_by_date.get(dt_str)
            if v and v[2] > 0:
                earn_expires_at = v[4]
            else:
                earn_expires_at = ""
                
            exp_v = expiring_vault_by_date.get(dt_str)
            if exp_v:
                exp_vault_id = exp_v[0]
                exp_is_extended = exp_v[5]
            else:
                exp_vault_id = None
                exp_is_extended = 0
                
            ledger.append({
                'date': dt_str,
                'earned_minutes': earned,
                'used_minutes': used,
                'expired_minutes': expired,
                'running_balance': running_balance,
                'earn_expires_at': earn_expires_at,
                'exp_vault_id': exp_vault_id,
                'exp_is_extended': exp_is_extended
            })
            
    ledger.sort(key=lambda x: x['date'], reverse=True)
    return ledger

def extend_overtime(user_id, vault_id, admin_pw, new_expires_at):
    conn = get_connection()
    c = conn.cursor()
    c.execute('SELECT password FROM admin_settings WHERE id = 1')
    row = c.fetchone()
    if not row or row[0] != admin_pw:
        conn.close()
        return False, "비밀번호가 일치하지 않습니다."
        
    c.execute('UPDATE overtime_vault SET expires_at = ?, is_extended = 1 WHERE id = ? AND user_id = ?', (new_expires_at, vault_id, user_id))
    conn.commit()
    conn.close()
    sync_vault(user_id)
    return True, "연장되었습니다."

def auto_fill_missing_days(user_id):
    pass

def auto_fill_missing_days(user_id, month_str):
    if not user_id: return 0
    from datetime import datetime, date, timedelta
    import calendar
    import sqlite3
    
    year, month = map(int, month_str.split('-'))
    num_days = calendar.monthrange(year, month)[1]
    
    start_date = date(year, month, 1)
    end_date = min(date(year, month, num_days), date.today() - timedelta(days=1))
    
    conn = get_connection()
    c = conn.cursor()
    c.execute('SELECT date FROM work_logs WHERE user_id = ? AND date LIKE ?', (user_id, month_str + '%'))
    existing_dates = set([row[0] for row in c.fetchall()])
    conn.close()
    
    curr = start_date
    count = 0
    while curr <= end_date:
        d_str = curr.strftime("%Y-%m-%d")
        if curr.weekday() < 5 and d_str not in existing_dates: # 평일이고 기록이 없으면
            save_work_log(user_id, d_str, 'normal', '08:30', '17:30', 0, 0, '')
            count += 1
        curr += timedelta(days=1)
    return count

def export_yearly_excel(year_str, target_user_id, filepath):
    import sqlite3
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    
    conn = get_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    if target_user_id == 'all':
        c.execute('SELECT id, name FROM users ORDER BY name')
        users = c.fetchall()
    else:
        c.execute('SELECT id, name FROM users WHERE id = ?', (target_user_id,))
        users = c.fetchall()
        
    wb = Workbook()
    if not users:
        wb.save(filepath)
        conn.close()
        return True
        
    first_sheet = True
    
    for u in users:
        u_id = u['id']
        u_name = u['name']
        
        c.execute('SELECT * FROM work_logs WHERE user_id = ? AND date LIKE ? ORDER BY date', (u_id, year_str + '%'))
        rows = c.fetchall()
        
        if first_sheet:
            ws = wb.active
            ws.title = u_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=u_name)
            
        def mins_to_hm(mins):
            if not mins or mins == 0: return "00:00"
            mins = int(mins)
            return f"{mins//60:02d}:{mins%60:02d}"

        def mins_to_hours(mins):
            if not mins or mins == 0: return None
            return round(int(mins) / 60, 2)

        def format_type(row):
            t_names = {
                "normal": "정상근무", "morning_half": "오전 반차", "afternoon_half": "오후 반차", 
                "holiday": "공휴일 / 특별휴가", "annual_leave": "연차", "public_leave": "공가", "sick_leave": "병가",
                "replacement_leave": "휴무", "public_leave_half": "반공가 + 반차"
            }
            base_name = t_names.get(row['type'], row['type'])
            desc = row['description']
            if row['type'] == 'holiday' and desc: return f"{base_name} - {desc}"
            if desc and row['type'] != 'holiday': return f"{base_name} ({desc})"
            return base_name

        expired_by_date_all = {}
        total_expired = 0
        for m in range(1, 13):
            stats = get_monthly_summary(u_id, f"{year_str}-{m:02d}")
            for k, v in stats.get('expired_by_date', {}).items():
                expired_by_date_all[k] = v
                total_expired += v
                
        stats_jan = get_monthly_summary(u_id, f"{year_str}-01")
        current_carry_over = stats_jan['carry_over']
        
        c.execute("SELECT MAX(date) as last_date FROM work_logs WHERE user_id = ? AND date LIKE ?", (u_id, year_str + '%'))
        last_date_row = c.fetchone()
        last_date = last_date_row['last_date'] if last_date_row and last_date_row['last_date'] else None
        
        headers = ['일자', '근무형태', '출근', '퇴근', '비근무시간', '휴게시간', '실근무시간', '발생연장근로', '사용연장근로', '소멸연장근로', '잔여연장근로']
        ws.append(headers)
        
        last_dt = f"{year_str}-00-00"
        actual_expired_total = 0
        
        for r in rows:
            dt_str = r['date']
            earned = r['overtime_earned'] or 0
            used = r['overtime_used'] or 0
            
            expired = sum(v for k, v in expired_by_date_all.items() if last_dt < k <= dt_str)
            actual_expired_total += expired
            last_dt = dt_str
            
            current_carry_over = current_carry_over + earned - used - expired
            
            ws.append([
                dt_str, format_type(r), r['clock_in'], r['clock_out'],
                mins_to_hm(r['non_work_time']), mins_to_hm(r['break_time']), mins_to_hm(r['work_time']),
                mins_to_hours(earned), mins_to_hours(used), mins_to_hours(expired), mins_to_hours(current_carry_over)
            ])
            
        try:
            yellow_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
            bold_font = Font(bold=True)
            center_align = Alignment(horizontal="center", vertical="center")
            
            text_red = Font(color="FF0000")
            text_blue = Font(color="0000FF")
            text_green = Font(color="008000")
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = bold_font
                cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                cell.alignment = center_align
                
            for r_idx, r in enumerate(rows, start=2):
                if (r['overtime_earned'] or 0) > 0:
                    for c_idx in range(1, len(headers) + 1):
                        ws.cell(row=r_idx, column=c_idx).fill = yellow_fill
                        
                t = r['type']
                if t in ['holiday', 'public_leave', 'annual_leave', 'sick_leave']:
                    ws.cell(row=r_idx, column=2).font = text_red
                elif t in ['replacement_leave']:
                    ws.cell(row=r_idx, column=2).font = text_blue
                elif t in ['morning_half', 'afternoon_half', 'public_leave_half']:
                    ws.cell(row=r_idx, column=2).font = text_green
                        
            # Yearly summary
            c.execute("SELECT SUM(overtime_earned) as e, SUM(overtime_used) as u FROM work_logs WHERE user_id = ? AND date LIKE ?", (u_id, year_str + '%'))
            yearly_stats = c.fetchone()
            y_earned = yearly_stats['e'] if yearly_stats['e'] else 0
            y_used = yearly_stats['u'] if yearly_stats['u'] else 0
            
            ws.cell(row=2, column=13, value="이월된 연장근로").font = bold_font
            ws.cell(row=2, column=14, value=round(stats_jan['carry_over'] / 60, 2))
            
            ws.cell(row=3, column=13, value="년도 총 발생 연장근로").font = bold_font
            ws.cell(row=3, column=14, value=round(y_earned / 60, 2))
            
            ws.cell(row=4, column=13, value="년도 총 사용 연장근로").font = bold_font
            ws.cell(row=4, column=14, value=round(y_used / 60, 2))
            
            red_font = Font(color="EF4444", bold=True)
            ws.cell(row=5, column=13, value="년도 총 소멸 연장근로").font = bold_font
            ws.cell(row=5, column=14, value=round(actual_expired_total / 60, 2)).font = red_font
                
            ws.cell(row=6, column=13, value="잔여 연장근로 시간").font = bold_font
            ws.cell(row=6, column=14, value=round(current_carry_over / 60, 2)).font = bold_font
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column].width = max_length + 2
        except Exception as style_e:
            pass

    wb.save(filepath)
    conn.close()
    return True
